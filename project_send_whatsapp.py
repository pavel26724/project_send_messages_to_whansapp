import pywhatkit
from openpyxl import load_workbook

book_client = load_workbook('client_base.xlsx')
sheet = book_client.active
a = []
for row in sheet.iter_rows(min_row=2):
    if str(row[0].value)[0] == '7' or str(row[0].value)[0] == '8':
        phone = '+7' + str(row[0].value)[1:]
    name = str(row[1].value).capitalize()
    type_of_work = str(row[2].value)
    male = str(row[3].value)
    send_or_not = str(row[4].value)
    if send_or_not == 'yes':
        a1 = [phone, name, type_of_work, male]
        a.append(a1)

m = open('text_man.txt', 'r', encoding='utf-8')
text_for_man = m.read()
w = open('text_women.txt', 'r', encoding='utf-8')
text_for_women = w.read()

for i in a:
    if i[3] == 'man':
        pywhatkit.sendwhatmsg_instantly(i[0], f'{i[1]}, {text_for_man}', 10, True, 3)
    else:
        pywhatkit.sendwhatmsg_instantly(i[0], f'{i[1]}, {text_for_women}', 10, True, 3)

m.close()
w.close()

